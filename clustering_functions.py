import os
import numpy as np
import math
from dtaidistance import dtw
from sklearn_extra.cluster import KMedoids
from sklearn.metrics import silhouette_score
from scipy.optimize import minimize
import matplotlib.pyplot as plt
import matplotlib as mpl
import numpy as np
import matplotlib.pyplot as plt
from sklearn.neighbors import LocalOutlierFactor
from openpyxl import Workbook, load_workbook
from fastdtw import fastdtw
from scipy.spatial.distance import euclidean

plt.rcParams["font.family"] =["Times New Roman"]
plt.rcParams["font.serif"] = ["Times New Roman"]
mpl.rcParams['figure.dpi'] = 600   

def bourdet_derivative(x=None,y=None,L=0.1,transform_x=False,transform_y=False,remove_end_points=False):
    # L is the smoothing factor 
    # x and y are arrays
    x=np.array(x)
    y=np.array(y)
    if transform_x:
        x=np.log(x)
    if transform_y:
        y=np.log(y)
        
    def forward_difference(yR,yC,xR,xC):
        a=(yR-yC)
        b=(xR-xC)
        return np.divide(a, b, out=np.zeros_like(a, dtype=float), where=b != 0)
    def backward_difference(yC,yL,xC,xL):
        a=(yC-yL)
        b=(xC-xL)
        return np.divide(a, b, out=np.zeros_like(a, dtype=float), where=b != 0)
    
    def higher_order_end_point(yL,yC,yR,xL,xC,xR,point=0):
        # Second-order forward difference for the first point
        if point==0:
            a=(-3*yL+4*yC-yR)
        else:
            a=(3*yR-4*yC+yL)
        b=(xR-xL)
        return np.divide(a, b, out=np.zeros_like(a, dtype=float), where=b != 0)
    
    b_diff=[]
    for i in range(len(x)):
        diff=x-x[i]
        eps=1e-4
        if i>0 and i<len(x)-1:        
            L_idx=np.where(diff<max(-L,min(diff)+eps))[0][np.argmax(x[diff<max(-L,min(diff)+eps)])]
            R_idx=np.where(diff>min(L,max(diff)-eps))[0][np.argmin(x[diff>min(L,max(diff)-eps)])]
            xL=x[L_idx]; yL=y[L_idx]
            xR=x[R_idx]; yR=y[R_idx]
            xC=x[i]; yC=y[i]
            fd=forward_difference(yR,yC,xR,xC)
            bd=backward_difference(yC,yL,xC,xL)
            diff=((xR-xC)*bd+(xC-xL)*fd)/(xR-xL)
        elif i==0:
            L_idx=0
            R_idx=np.where(diff>min(L,max(diff)-eps))[0][np.argmin(x[diff>min(L,max(diff)-eps)])]
            xR=x[R_idx]; yR=y[R_idx]
            xC=x[i]; yC=y[i]            
            diff=forward_difference(yR,yC,xR,xC)
        elif i==len(x)-1:
            L_idx=np.where(diff<max(-L,min(diff)+eps))[0][np.argmax(x[diff<max(-L,min(diff)+eps)])]
            R_idx=len(x)-1
            xL=x[L_idx]; yL=y[L_idx]
            xC=x[i]; yC=y[i]
            diff=backward_difference(yC,yL,xC,xL)

        #print(fd,bd,cd,i)       
        b_diff.append(diff)

    #if remove_end_points:
        #b_diff[0]=b_diff[1]
        #b_diff[-1]=b_diff[-2]

    return np.array(b_diff)

def bourdet_derivative_old(x=None, y=None, L=0.2,transform_x=False,transform_y=False,):
    """
    Calculate the Bourdet derivative of pressure with respect to ln(time)
    for well-test analysis.
    
    The derivative at each point is computed using one point to the left 
    and one to the right that are at least L log-cycles away. This approach
    both computes and smooths the derivative.
    
    Parameters:
        t (array-like): Time data (must be > 0).
        p (array-like): Corresponding pressure data.
        L (float): Log-cycle window for smoothing (typical values range from 0.01 to 0.2).
    
    Returns:
        dP_dlnT (np.ndarray): The smoothed derivative dp/d(ln(t)) at each time point.
    
    The derivative at an interior point is estimated as:
        dp/dln(t) = 0.5 * [ (p[i] - p[left]) / (ln(t[i]) - ln(t[left])) 
                          + (p[right] - p[i]) / (ln(t[right]) - ln(t[i])) ]
    For the first and last points, a one-sided difference is used.
    
    References:
        Bourdet, D., Ayoub, J.A., & Pirard, Y.M. (1989). Use of Pressure Derivative in
        Well-Test Interpretation. SPE Formation Evaluation, June 1989.
    """
    t = np.asarray(x)
    p = np.asarray(y)
    n = len(t)
    dP_dlnT = np.zeros(n)
    ln_t=t
    if transform_x:
       ln_t = np.log(t)
    
    for i in range(n):
        # Find the left index j where the log difference is at least L.
        j = i
        while j > 0 and (ln_t[i] - ln_t[j-1]) < L:
            j -= 1
        # Ensure j doesn't fall below index 0.
        j = max(j, 0)
        
        # Find the right index k where the log difference is at least L.
        k = i
        while k < n - 1 and (ln_t[k+1] - ln_t[i]) < L:
            k += 1
        k = min(k, n - 1)
        
        if i == 0:
            # For the first point, use a forward difference.
            dt = ln_t[k] - ln_t[i]
            dP_dlnT[i] = (p[k] - p[i]) / dt if dt != 0 else 0.0
        elif i == n - 1:
            # For the last point, use a backward difference.
            dt = ln_t[i] - ln_t[j]
            dP_dlnT[i] = (p[i] - p[j]) / dt if dt != 0 else 0.0
        else:
            # Use a central difference by averaging left and right estimates.
            dt_left = ln_t[i] - ln_t[j]
            dt_right = ln_t[k] - ln_t[i]
            d_left = (p[i] - p[j]) / dt_left if dt_left != 0 else 0.0
            d_right = (p[k] - p[i]) / dt_right if dt_right != 0 else 0.0
            # dP_dlnT[i] = 0.5 * (d_left + d_right)
            dP_dlnT[i] = ((d_left*dt_right) +(d_right*dt_left)) /(dt_left+dt_right)
    
    return dP_dlnT

def compute_angle(seq1, seq2):
    # Convert sequences to numpy arrays
    vec1 = np.array(seq1)
    vec2 = np.array(seq2)
    
    # Pad with zeros for unequal lengths of vectors
    if len(vec1)!=len(vec2):
        max_length = max(len(vec1), len(vec2))
        vec1 = vec1 + [0] * (max_length - len(vec1))
        vec2 = vec2 + [0] * (max_length - len(vec2))
    
    # Compute dot product and magnitudes
    dot_product = np.dot(vec1, vec2)
    norm_vec1 = np.linalg.norm(vec1)
    norm_vec2 = np.linalg.norm(vec2)
    
    # Compute cosine similarity
    cosine_similarity = dot_product / (norm_vec1 * norm_vec2)
    
    # Clip the cosine similarity to avoid numerical issues with arccos
    cosine_similarity = np.clip(cosine_similarity, -1.0, 1.0)
    
    # Compute the angle in radians
    angle_radians = np.arccos(cosine_similarity)
    
    # Convert the angle to degrees
    angle_degrees = np.degrees(angle_radians)
    
    return angle_degrees

# Example sequences
sequence1 = [4, 7, 6, 2, 3]
sequence2 = [2, 5, 6, 8, 9]

angle = compute_angle(sequence1, sequence2)
print(f"Angle between the sequences: {angle:.2f} degrees")

# Compute the distance matrix
def compute_cosine_similarity_matrix(subsequences):
    num_series = len(subsequences)
    distance_matrix = np.zeros((num_series, num_series))

    for i in range(num_series):
        for j in range(i, num_series):
            dist = compute_angle(subsequences[i], subsequences[j])
            distance_matrix[i, j] = dist
            distance_matrix[j, i] = dist
    return distance_matrix

def compute_distance_matrix(subsequences):
    num_series = len(subsequences)
    distance_matrix = np.zeros((num_series, num_series))

    for i in range(num_series):
        for j in range(i, num_series):
            if j<=(i+1):
                dist = dtw.distance(subsequences[i], subsequences[j])
            else:
                dist=np.zeros([])
            distance_matrix[i, j] = dist
            distance_matrix[j, i] = dist
    return distance_matrix

def compute_distance_matrix_ND(time_series):
    # Initialize distance matrix
    n = len(time_series)
    distance_matrix = np.zeros((n, n))
    
    # Compute pairwise DTW distances
    for i in range(n):
        for j in range(i, n):  # Upper triangle to exploit symmetry
            distance, _ = fastdtw(time_series[i], time_series[j], dist=euclidean)
            distance_matrix[i, j] = distance
            distance_matrix[j, i] = distance  # Mirror to lower triangle
    
    return distance_matrix

def dot_product_padding(a, b):
    max_length = max(len(a), len(b))
    a_padded = a + [0] * (max_length - len(a))
    b_padded = b + [0] * (max_length - len(b))
    return sum(a_padded[i] * b_padded[i] for i in range(max_length))

# for row in dtw_dist_matrix:
#     y_idx=np.where(np.all(dtw_dist_matrix == row, axis=1))[0][0]
#     dmat_row=[row[i+1]-row[i] for i in range(len(row)-1)]
#     dmin=min(row[1:][row[1:]>0.])
#     for dc in dmat_row:
#         x_idx=dmat_row.index(dc)
#         if abs(dc)>0.1:
#             dtw_dist_matrix[y_idx,x_idx]=dmin

# Evaluate the shiloutte loss for a given number of k-Medoids clusters
def evaluate_silhouette(params,time_series):
    
    # Params is a list of initial parameters
    k=int(params[0])
    window_size = int(params[1])
    # step_size = int(params[2])
    # Ensure the parameters are within a valid range
    if window_size <= 0 or window_size >= len(time_series[0]):
        return np.inf  # Penalize invalid parameters

    new_subsequences=[]
    # Create subsequences
    for seq in time_series:
        norm_feature=min_max_scaler(seq,limits=[-1,1])
        subsequence=create_subsequences(norm_feature, window_size, window_size)
        new_subsequences.append(subsequence)

    subsequences_i=[interleave_multiple_arrays(*arg) for arg in zip(*new_subsequences)]
    dtw_dist_matrix=dtw.distance_matrix_fast(subsequences_i)
    dtw_dist_matrix=(dtw_dist_matrix)#+dtw_dist_matrix+cosine_similarity

    # Compute pairwise DTW distance matrix
    # n = len(subsequences)
    # distance_matrix = np.zeros((n, n))

    # for i in range(n):
    #     for j in range(i + 1, n):
    #         distance = dtw.distance(subsequences[i], subsequences[j])
    #         distance_matrix[i, j] = distance
    #         distance_matrix[j, i] = distance

    # Apply K-medoids clustering
    kmedoids = KMedoids(n_clusters=k, metric='precomputed', random_state=42)
    labels = kmedoids.fit_predict(dtw_dist_matrix)
    
    # Compute silhouette score
    score = silhouette_score(dtw_dist_matrix, labels, metric='precomputed')

    return -score

def lof_outlier_removal(x_values, y_values, n_neighbors=20, contamination=0.1):
    """
    Performs LOF outlier removal on 2D data defined by x_values and y_values.
    For each outlier detected, replaces the point with the closest inlier.
    
    Parameters:
      x_values (list): List of x coordinates.
      y_values (list): List of y coordinates.
      n_neighbors (int): Number of neighbors to use for LOF.
      contamination (float): The expected proportion of outliers.
      
    Returns:
      original_data (np.array): The original 2D data as an array of shape (n_samples, 2).
      labels (np.array): LOF labels (1 for inliers, -1 for outliers).
      processed_data (np.array): The modified data after replacing outliers.
    """
    # Combine x and y into a 2D numpy array (each row is an (x, y) pair)
    original_data = np.array(list(zip(x_values, y_values)))
    
    # Apply LOF to detect outliers
    lof = LocalOutlierFactor(n_neighbors=n_neighbors, contamination=contamination)
    labels = lof.fit_predict(original_data)
    
    # Copy the original data to modify outliers
    processed_data = original_data.copy()
    
    # For each outlier, find the closest inlier and replace the outlier point
    for i in range(len(original_data)):
        if labels[i] == -1:
            min_dist = float('inf')
            closest_point = None
            for j in range(len(original_data)):
                if labels[j] == 1:
                    dist = np.linalg.norm(original_data[i] - original_data[j])
                    if dist < min_dist:
                        min_dist = dist
                        closest_point = original_data[j]
            if closest_point is not None:
                processed_data[i] = closest_point
                
    return original_data, labels, processed_data

# Normalize and combine (with safety checks)
import numpy as np
from dtaidistance import dtw
from sklearn_extra.cluster import KMedoids

def hybrid_distance_with_penalty(segments, alpha=0.5, gamma=0.3):
    """
    Hybrid distance with positional penalty to prevent cluster repetition.
    :param segments: List of N-dimensional segments (shape: [n_segments, window_size, n_features])
    :param alpha: Weight for DTW vs Euclidean (0.5 = equal weighting)
    :param gamma: Strength of positional penalty (0 = no penalty, 0.5 = strong penalty)
    """
    n = len(segments)
    dtw_dist = np.zeros((n, n))
    eucl_dist = np.zeros((n, n))
    pos_penalty = np.zeros((n, n))

    # Compute DTW, Euclidean, and positional penalty
    for i in range(n):
        for j in range(n):
            # DTW distance for N-dimensional segments
            dtw_dist[i, j] = dtw.distance(segments[i], segments[j], use_ndim=True)
            
            # Euclidean distance between centroids
            centroid_i = np.mean(segments[i], axis=0)
            centroid_j = np.mean(segments[j], axis=0)
            eucl_dist[i, j] = np.linalg.norm(centroid_i - centroid_j)
            
            # Positional penalty (temporal/spatial distance between segments)
            pos_penalty[i, j] = abs(i - j)  # Linear penalty based on segment index

    # Safe normalization
    epsilon = 1e-8
    dtw_norm = dtw_dist / (np.max(dtw_dist) + epsilon)
    eucl_norm = eucl_dist / (np.max(eucl_dist) + epsilon)
    pos_norm = pos_penalty / (np.max(pos_penalty) + epsilon)

    # Combine with weights
    hybrid = alpha * dtw_norm + (1 - alpha) * eucl_norm
    total_distance = hybrid + gamma * pos_norm  # Add positional penalty

    return total_distance

def hybrid_distance_eucl_cos(segments, alpha=0.5, gamma=0.3):
    """
    Combines Euclidean distance and Cosine similarity with positional penalty.
    :param segments: List of N-dimensional segments (shape: [n_segments, window_size, n_features])
    :param alpha: Weight between Euclidean (alpha) and Cosine (1-alpha)
    :param gamma: Strength of positional penalty
    """
    n = len(segments)
    eucl_dist = np.zeros((n, n))
    cos_dist = np.zeros((n, n))
    pos_penalty = np.zeros((n, n))

    # Pre-flatten segments for vector operations

    flattened_segments = [seg.flatten() for seg in segments]

    # Pairwise computations
    for i in range(n):
        for j in range(n):
            # Euclidean distance between flattened vectors
            eucl_dist[i,j] = np.linalg.norm(flattened_segments[i] - flattened_segments[j])
            
            # Cosine distance (1 - cosine similarity)
            vec_i = flattened_segments[i]
            vec_j = flattened_segments[j]
            norm_i = np.linalg.norm(vec_i)
            norm_j = np.linalg.norm(vec_j)
            if norm_i == 0 or norm_j == 0:
                cos_dist[i,j] = 1  # Handle zero vectors
            else:
                cos_dist[i,j] = 1 - np.dot(vec_i, vec_j) / (norm_i * norm_j + 1e-8)
            
            # Positional penalty (segment index difference)
            pos_penalty[i,j] = abs(i - j)

    # Normalize with epsilon for stability
    epsilon = 1e-8
    eucl_norm = eucl_dist / (np.max(eucl_dist) + epsilon)
    cos_norm = cos_dist / (np.max(cos_dist) + epsilon)
    pos_norm = pos_penalty / (np.max(pos_penalty) + epsilon)

    # Combine distances and penalty
    hybrid = alpha * eucl_norm + (1 - alpha) * cos_norm
    total_distance = hybrid + gamma * pos_norm

    return total_distance
def interleave_multiple_arrays(*arrays):
    """
    Interleave multiple 1D numpy arrays element by element.
    :param arrays: Multiple numpy arrays (passed as arguments)
    :return: Interleaved numpy array
    """
    # Ensure that all arrays are numpy arrays and check they have the same length
    arrays = [np.array(a) for a in arrays]
    lengths = [len(a) for a in arrays]
    min_length = min(lengths)

    # Interleave the common part
    interleaved = np.empty(min_length * len(arrays), dtype=arrays[0].dtype)
    for i in range(len(arrays)):
        interleaved[i::len(arrays)] = arrays[i][:min_length]

    # Append the remaining elements from any longer arrays
    for i, array in enumerate(arrays):
        if len(array) > min_length:
            interleaved = np.concatenate((interleaved, array[min_length:]))

    return interleaved

def pair_multiple(*lists):
    """
    Given several lists of lists (e.g., A, B, C, ...),
    returns a list where each element is built by zipping together
    the corresponding sublists from each input list.
    
    For example:
      A = [[x1, x2, x3, x4], [x5, x6, x7, x8]]
      B = [[y1, y2, y3, y4], [y5, y6, y7, y8]]
      C = [[z1, z2, z3, z4], [z5, z6, z7, z8]]
      
    pair_multiple(A, B, C)
    
    returns:
      [
        [[x1, y1, z1], [x2, y2, z2], [x3, y3, z3], [x4, y4, z4]],
        [[x5, y5, z5], [x6, y6, z6], [x7, y7, z7], [x8, y8, z8]]
      ]
    """
    # zip(*lists) pairs the corresponding sublists from each input list.
    # Then, for each group of sublists, zip(*group) pairs up the corresponding
    # elements inside the sublists.
    a = [
        [list(items) for items in zip(*group)]
        for group in zip(*lists)
    ]
    return [np.array(a[i]) for i in range(len(a))]

def nearest_step_function(x, c):
    """
    Converts input x to the nearest value in list c, creating a step function.
    
    Parameters:
    - x: A scalar or array of input values.
    - c: A list of target values to which x will be mapped.
    
    Returns:
    - A scalar or array of the same shape as x, with each element mapped to the nearest value in c.
    """
    # Convert x to a numpy array if it's not already
    x = np.asarray(x)
    
    # Compute the absolute differences between x and each value in c
    differences = np.abs(x[..., np.newaxis] - np.array(c))
    
    # Find the indices of the minimum differences
    nearest_indices = np.argmin(differences, axis=-1)
    
    # Map each input value to the nearest value in c
    return np.array(c)[nearest_indices]

def min_max_scaler(x,limits=[-1,1]):
    x_min=np.min(x,axis=0)
    x_max=np.max(x,axis=0)
    return ((x-x_min)/(x_max-x_min))*(limits[1]-limits[0])+limits[0]

def create_subsequences(time_series,window_size,step_size):
    #subsequences = np.array([key_seq[i:i+subseq_len] for i in range(len(key_seq) - subseq_len + 1)],dtype=np.double)
    return np.array([time_series[i:i + window_size] for i in range(0, len(time_series) - window_size + 1, step_size)],dtype=np.double)

def map_labels_to_increasing_values(arr):
    """
    Maps the unique labels in the input array to increasing integer values based on their first occurrence.

    Parameters:
    arr (numpy.ndarray): Input array with labels to be mapped.

    Returns:
    numpy.ndarray: Array with labels mapped to increasing integer values.
    """
    # Ensure the input is a NumPy array
    arr = np.asarray(arr)
    
    # Find the unique labels in the order they appear
    unique_labels, first_occurrences = np.unique(arr, return_index=True)
    sorted_indices = np.argsort(first_occurrences)
    unique_labels = unique_labels[sorted_indices]
    
    # Create a mapping from original labels to new increasing values
    label_mapping = {label: idx for idx, label in enumerate(unique_labels)}
    
    # Apply the mapping to the original array
    mapped_arr = np.vectorize(label_mapping.get)(arr)
    
    return mapped_arr

def copy_arrays_to_excel(arrays, file_path, sheet_name='x'):
    """
    Copies a list of 1D arrays into an Excel file.
    
    If the Excel file exists and a sheet with 'sheet_name' exists,
    a new sheet will be created using the name 'sheet_name_new'. If that sheet
    also exists, an incrementing counter is appended until a unique name is found.
    
    Each array in 'arrays' is written as a column.
    
    Parameters:
        arrays (list): List of 1D arrays (or lists) to write.
        file_path (str): Path to the Excel file.
        sheet_name (str): Preferred sheet name. Defaults to 'x'.
    """
    # Load workbook if it exists; otherwise, create a new one.
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
    
    # Determine a unique sheet name.
    if sheet_name in wb.sheetnames:
        new_sheet_name = sheet_name + '_new'
        counter = 1
        # If even the new name exists, keep appending a counter until unique.
        while new_sheet_name in wb.sheetnames:
            counter += 1
            new_sheet_name = f"{sheet_name}_new{counter}"
    else:
        new_sheet_name = sheet_name

    # If the workbook is new and only the default sheet exists,
    # you can rename it instead of creating a new one.
    if len(wb.sheetnames) == 1 and wb.active.title == "Sheet" and new_sheet_name != "Sheet":
        ws = wb.active
        ws.title = new_sheet_name
    else:
        ws = wb.create_sheet(title=new_sheet_name)
    
    # Write each array as a column.
    for col_idx, arr in enumerate(arrays, start=1):
        for row_idx, value in enumerate(arr, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)

def k_medoids_plot(key_seq,lndt,grad_3,lndp_dlndt,params=[None,None,None]):
    # Use normalized inputs
    k=int(params[0])
    window_size = int(params[1])
    step_size = window_size#int(params[2])
    key_seq=min_max_scaler(key_seq,limits=[-1,1])
    lndt=min_max_scaler(lndt,limits=[-1,1])
    grad_3=min_max_scaler(grad_3,limits=[-1,1])
    lndp_dlndt=min_max_scaler(lndp_dlndt,limits=[-1,1])           # Used for plotting
    
    subsequences=create_subsequences(key_seq, window_size, step_size)
    subsequences_idx=create_subsequences(lndt, window_size, step_size)
    subsequences_aux1=create_subsequences(grad_3, window_size, step_size)
    subsequences_lndp_dlndt=create_subsequences(lndp_dlndt, window_size, step_size) #Used for plotting
    sub_list=subsequences,subsequences_idx,subsequences_aux1,
    
    subsequences_i=[interleave_multiple_arrays(*args) for args in zip(*sub_list)]
    #subsequences_i=[cf.interleave_multiple_arrays(i,j,) for i,j in zip(subsequences,subsequences_idx,)]
    dtw_dist_matrix=dtw.distance_matrix_fast(subsequences_i)  

    # Apply k-medoids clustering
    kmedoids = KMedoids(n_clusters=k, metric='precomputed', method='pam',random_state=42)
    kmedoids.fit(dtw_dist_matrix)

    medoid_indices = kmedoids.medoid_indices_
    labels = kmedoids.labels_
    # Step 4: Plot the clusters
    colors = ['r', 'g', 'b', 'c', 'm', 'y']
    plt.figure(figsize=(10, 6))
    fig, ax = plt.subplots(2,1,dpi=1200,figsize=(6,4))
    fig.subplots_adjust(left=0.02, bottom=0.06, right=0.95, top=0.95, hspace=0.25, wspace=0.5)  

    for idx, seq in enumerate(subsequences):
        color_n0 = colors[labels[idx] % len(colors)]
        label_n0=labels[idx]
        
        # Get the next label in the subsequence
        if idx>0:
            label_n_1=labels[idx-1]
        else:
            label_n_1=label_n0
    
        if (label_n0==label_n_1) or (step_size>=window_size):
            # No overlapping
            seq_n=seq
            seq_dp_n=subsequences_lndp_dlndt[idx]
            seq_idx_n=subsequences_idx[idx]
        else:
            # Overlapping
            # Get the max value of the n-1 x-list 
            max_n_1=max(subsequences_idx[idx-1])
            min_n0=min(subsequences_idx[idx])
            mid_val=(min_n0+max_n_1)/2
            
            # Get x-y values greater than mid value
            seq_n=seq[seq>mid_val]
            seq_dp_n=subsequences_lndp_dlndt[idx][seq>mid_val]
            seq_idx_n=subsequences_idx[idx][seq>mid_val]
            lbl=f'Cluster {label_n0 + 1}'
                       
        ax[0].plot(seq_idx_n,seq_n,marker='o',markerfacecolor='none',markeredgecolor=color_n0, markeredgewidth=1, color=color_n0, label='')
        ax[1].plot(seq_idx_n,seq_dp_n, marker='s',markerfacecolor='none',markeredgecolor=color_n0, markeredgewidth=1,color=color_n0, label='')


    # Highlight the medoids
    for medoid_idx in medoid_indices:
        medoid_seq = subsequences[medoid_idx]
        medoid_seq_idx=subsequences_idx[medoid_idx]
        label=f'Medoid_Cluster_{list(medoid_indices).index(medoid_idx)+1}_subseq_idx_{medoid_idx+1}'
        ax[0].plot(medoid_seq_idx,medoid_seq, marker='*', color='k', markersize=15, label=label)
    
    fig.suptitle('1D Sequence Clustering with DTW and k-Medoids',y=1.01)
    ax[1].set_xlabel('lndt')
    ax[1].set_ylabel('Normalized Pressure Difference')
    plt.legend()
    plt.show()

    return